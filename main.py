from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


now1 = datetime.now()
print(now1)
now1 = str(now1).split(' ')
data_inicio = now1[0]
hora_inicio = now1[1]
hora_inicio = hora_inicio.split('.')
hora_inicio = hora_inicio[0]
print('DATA INICIO: '+data_inicio)
print('HORA INICIO: '+hora_inicio)


servico = Service(ChromeDriverManager().install())

driver = webdriver.Chrome(service=servico)

driver.maximize_window()

driver.get("https://tipminer.com/blaze/double")

black_to_black=0
black_to_white=0
black_to_red=0
white_to_white=0
white_to_red=0
white_to_black=0
red_to_red=0
red_to_white=0
red_to_black=0

roll_black='roll black '
roll_red='roll red '
roll_white='roll white '

peido = True
cont=0
quant=20
sequencia=[]

while peido==True and cont<quant:

    cont = cont+1

    cor1 = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[2]/div[1]").get_attribute("class")
    cor2 = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
    print('COR1: '+str(cor1))
    print('COR2: '+str(cor2))

    if cor1 == roll_black and cor2 == roll_black:
        black_to_black = black_to_black+1
        now2 = datetime.now()
        sequencia.append('black_to_black-'+str(now2))
        print('black_to_black: '+str(now2))
        time.sleep(30)
    
    elif cor1 == roll_black and cor2 == roll_white:
        black_to_white = black_to_white+1
        now2 = datetime.now()
        sequencia.append('black_to_white-'+str(now2))
        print('black_to_white: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_black and cor2 == roll_red:
        black_to_red = black_to_red+1
        now2 = datetime.now()
        sequencia.append('black_to_red-'+str(now2))
        print('black_to_red: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_white and cor2 == roll_white:
        white_to_white = white_to_white+1
        now2 = datetime.now()
        sequencia.append('white_to_white-'+str(now2))
        print('white_to_white: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_white and cor2 == roll_red:
        white_to_red = white_to_red+1
        now2 = datetime.now()
        sequencia.append('white_to_red-'+str(now2))
        print('white_to_red: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_white and cor2 == roll_black:
        white_to_black = white_to_black+1
        now2 = datetime.now()
        sequencia.append('white_to_black-'+str(now2))
        print('white_to_black: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_red and cor2 == roll_red:
        red_to_red = red_to_red+1
        now2 = datetime.now()
        sequencia.append('red_to_red-'+str(now2))
        print('red_to_red: '+str(now2))
        time.sleep(30)

    elif cor1 == roll_red and cor2 == roll_white:
        red_to_white = red_to_white+1
        now2 = datetime.now()
        sequencia.append('red_to_white-'+str(now2))
        print('red_to_white: '+str(now2))
        time.sleep(30)

    elif cor1== roll_red and cor2 == roll_black:
        red_to_black = red_to_black+1
        now2 = datetime.now()
        sequencia.append('red_to_black-'+str(now2))
        print('red_to_black: '+str(now2))
        time.sleep(30)

    print('CONT: '+str(cont))

print('BLACK TO BLACK: '+str(black_to_black))
print('BLACK TO WHITE: '+str(black_to_white))
print('BLACK TO RED: '+str(black_to_red))
print('WHITE TO WHITE: '+str(white_to_white))
print('WHITE TO RED: '+str(white_to_red))
print('WHITE TO BLACK: '+str(white_to_black))
print('RED TO RED: '+str(red_to_red))
print('RED TO BLACK: '+str(red_to_black))
print('RED TO WHITE: '+str(red_to_white))

now2 = str(now2).split(' ')
data_final = now2[0]
hora_final = now2[1]
hora_final = hora_final.split('.')
hora_final = hora_final[0]
hora_final = hora_final.replace(":","-")
hora_inicio = hora_inicio.replace(":","-")

lista_combinações = ['BLACK TO BLACK','BLACK TO WHITE','BLACK TO RED','WHITE TO WHITE','WHITE TO RED','WHITE TO BLACK',
    'RED TO RED','RED TO BLACK','RED TO WHITE']

lista_combinacoes_resultados = [black_to_black,black_to_white,black_to_red,white_to_white,white_to_red,white_to_black,red_to_red,
    red_to_black,red_to_white]


wb = Workbook()
ws1 = wb.active
ws1.title = data_inicio+'_'+hora_inicio+'_'+hora_final

for col in range(1,4):
    if col == 1:
        for row in range(1,10):
            letter = get_column_letter(col)
            ws1[letter + str(row)] = lista_combinações[row-1]
    elif col == 2:
        for row in range(1,10):
            letter = get_column_letter(col)
            ws1[letter + str(row)] = lista_combinacoes_resultados[row-1]
    elif col == 3:
        for row in range(1,11):
            letter = get_column_letter(col)
            ws1[letter + str(row)] = sequencia[row-1]

ws2 = wb.create_sheet(title="Ok")
ws2["C1"] = "OK"


wb.save('('+str(quant)+')Relatorio_Robo_'+data_inicio+'_'+hora_inicio+'_'+hora_final+'.xlsx')
        

    





