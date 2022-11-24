from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import email, smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import telebot 
from telethon.sync import TelegramClient 
from telethon.tl.types import InputPeerUser, InputPeerChannel 
from telethon import TelegramClient, sync, events 
import requests

now1 = datetime.now()
print(now1)
now1 = str(now1).split(' ')
data_inicio = now1[0]
hora_inicio = now1[1]
hora_inicio = hora_inicio.split('.')
hora_inicio = hora_inicio[0]
print('DATA INICIO: '+data_inicio)
print('HORA INICIO: '+hora_inicio)

servico = Service(ChromeDriverManager().install())

driver = webdriver.Chrome(service=servico)

driver.maximize_window()

roll_black='roll black '
roll_red='roll red '
roll_white='roll white '

quant=5
sequencia=[]

def last_chat_id(token):
    try:
        url = "https://api.telegram.org/bot{}/getUpdates".format(token)
        response = requests.get(url)
        if response.status_code == 200:
            json_msg = response.json()
            for json_result in reversed(json_msg['result']):
                message_keys = json_result['message'].keys()
                if ('new_chat_member' in message_keys) or ('group_chat_created' in message_keys):
                    return json_result['message']['chat']['id']
            print('Nenhum grupo encontrado')
        else:
            print('A resposta falhou, código de status: {}'.format(response.status_code))
    except Exception as e:
        print("Erro no getUpdates:", e)

# enviar mensagens utilizando o bot para um chat específico
def send_message(token, chat_id, message):
    try:
        data = {"chat_id": chat_id, "text": message}
        url = "https://api.telegram.org/bot{}/sendMessage".format(token)
        requests.post(url, data)
    except Exception as e:
        print("Erro no sendMessage:", e)

while True:
    black_to_black=0
    black_to_white=0
    black_to_red=0
    white_to_white=0
    white_to_red=0
    white_to_black=0
    red_to_red=0
    red_to_white=0
    red_to_black=0
    cont=0
    while cont<quant:

        cont = cont+1

        driver.get("https://tipminer.com/blaze/double")

        cor1 = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[2]/div[1]").get_attribute("class")
        cor2 = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
        print('COR1: '+str(cor1))
        print('COR2: '+str(cor2))

        if cor1 == roll_black and cor2 == roll_black:
            black_to_black = black_to_black+1
            now2 = datetime.now()
            sequencia.append('black_to_black-'+str(now2))
            print('black_to_black: '+str(now2))
            time.sleep(30)
        
        elif cor1 == roll_black and cor2 == roll_white:
            black_to_white = black_to_white+1
            now2 = datetime.now()
            sequencia.append('black_to_white-'+str(now2))
            print('black_to_white: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_black and cor2 == roll_red:
            black_to_red = black_to_red+1
            now2 = datetime.now()
            sequencia.append('black_to_red-'+str(now2))
            print('black_to_red: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_white and cor2 == roll_white:
            white_to_white = white_to_white+1
            now2 = datetime.now()
            sequencia.append('white_to_white-'+str(now2))
            print('white_to_white: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_white and cor2 == roll_red:
            white_to_red = white_to_red+1
            now2 = datetime.now()
            sequencia.append('white_to_red-'+str(now2))
            print('white_to_red: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_white and cor2 == roll_black:
            white_to_black = white_to_black+1
            now2 = datetime.now()
            sequencia.append('white_to_black-'+str(now2))
            print('white_to_black: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_red and cor2 == roll_red:
            red_to_red = red_to_red+1
            now2 = datetime.now()
            sequencia.append('red_to_red-'+str(now2))
            print('red_to_red: '+str(now2))
            time.sleep(30)

        elif cor1 == roll_red and cor2 == roll_white:
            red_to_white = red_to_white+1
            now2 = datetime.now()
            sequencia.append('red_to_white-'+str(now2))
            print('red_to_white: '+str(now2))
            time.sleep(30)

        elif cor1== roll_red and cor2 == roll_black:
            red_to_black = red_to_black+1
            now2 = datetime.now()
            sequencia.append('red_to_black-'+str(now2))
            print('red_to_black: '+str(now2))
            time.sleep(30)

        print('CONT: '+str(cont))

    maior = 0
    mais_de_um_maior=False

    lista_combinacoes_resultados = [black_to_black,black_to_white,black_to_red,white_to_white,white_to_red,white_to_black,red_to_red,
        red_to_black,red_to_white]

    # lista_combinacoes_resultados = [0,1,2,1,1,0,3,
    #     2,0]

    print('LISTA COMBINAÇÕES: '+str(lista_combinacoes_resultados))

    for x in range(0,len(lista_combinacoes_resultados)):
        print(lista_combinacoes_resultados[x])
        if lista_combinacoes_resultados[x] > maior:
            mais_de_um_maior=False
            maior=0
            maior = lista_combinacoes_resultados[x]
            if x == 0:
                opcao_maior='BLACK TO BLACK'
            elif x == 1:
                opcao_maior='BLACK TO WHITE'
            elif x == 2:
                opcao_maior='BLACK TO RED'
            elif x == 3:
                opcao_maior='WHITE TO WHITE'
            elif x == 4:
                opcao_maior='WHITE TO RED'
            elif x == 5:
                opcao_maior='WHITE TO BLACK'
            elif x == 6:
                opcao_maior='RED TO RED'
            elif x == 7:
                opcao_maior='RED TO BLACK'
            elif x == 8:
                opcao_maior='RED TO WHITE'

        elif lista_combinacoes_resultados[x] == maior and maior!=0:
            mais_de_um_maior=True

    maior = opcao_maior+' - '+str(maior)+' vezes'
    if mais_de_um_maior==True:
        print('MAIS DE UM MAIOR')
        continue


    print('BLACK TO BLACK: '+str(black_to_black))
    print('BLACK TO WHITE: '+str(black_to_white))
    print('BLACK TO RED: '+str(black_to_red))
    print('WHITE TO WHITE: '+str(white_to_white))
    print('WHITE TO RED: '+str(white_to_red))
    print('WHITE TO BLACK: '+str(white_to_black))
    print('RED TO RED: '+str(red_to_red))
    print('RED TO BLACK: '+str(red_to_black))
    print('RED TO WHITE: '+str(red_to_white))

    now2 = str(now2).split(' ')
    data_final = now2[0]
    hora_final = now2[1]
    hora_final = hora_final.split('.')
    hora_final = hora_final[0]
    hora_final = hora_final.replace(":","-")
    hora_inicio = hora_inicio.replace(":","-")

    lista_combinações = ['BLACK TO BLACK','BLACK TO WHITE','BLACK TO RED','WHITE TO WHITE','WHITE TO RED','WHITE TO BLACK',
        'RED TO RED','RED TO BLACK','RED TO WHITE']


    wb = Workbook()
    ws1 = wb.active
    ws1.title = data_inicio+'_'+hora_inicio+'_'+hora_final

    if mais_de_um_maior != True:
        while len(driver.find_elements(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]")) <1:
            print('esperando aparecer quadradinho double')
            time.sleep(1)
        if 'BLACK TO BLACK' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_black:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_black:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'BLACK TO WHITE' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_black:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_white:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'BLACK TO RED' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_black:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_red:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'WHITE TO WHITE' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_white:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_white:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'WHITE TO RED' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_white:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_red:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'WHITE TO BLACK' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_white:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_black:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'RED TO RED' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_red:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_red:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'RED TO BLACK' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_red:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_black:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)
        if 'RED TO WHITE' in maior:
            while True:
                print('esperando gatilho...')
                proximo_depois_do_ultimo = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                print('Proximo depois do ultimo: '+ str(proximo_depois_do_ultimo))
                if proximo_depois_do_ultimo == roll_red:
                    time.sleep(30)
                    temp = driver.find_element(By.XPATH,"/html/body/main/div/div/div[2]/div[2]/div[2]/div/div/div[1]/div[1]").get_attribute("class")
                    print('temp: '+ str(temp))
                    if temp == roll_white:
                        result='ACERTOU -'+temp
                    else:
                        result='ERROU -'+temp
                    break
                else:
                    time.sleep(30)



    for col in range(1,6):
        if col == 1:
            for row in range(1,10):
                letter = get_column_letter(col)
                ws1[letter + str(row)] = lista_combinações[row-1]
        elif col == 2:
            for row in range(1,10):
                letter = get_column_letter(col)
                ws1[letter + str(row)] = lista_combinacoes_resultados[row-1]
        elif col == 3:
            for row in range(1,quant+1):
                letter = get_column_letter(col)
                ws1[letter + str(row)] = sequencia[row-1]
        elif col == 4:
            if mais_de_um_maior != True:
                for row in range(1,2):
                    letter = get_column_letter(col)
                    ws1[letter + str(row)] = maior
            else:
                for row in range(1,2):
                    letter = get_column_letter(col)
                    ws1[letter + str(row)] = 'Mais de 1 resultado maior'
        elif col == 5 and mais_de_um_maior != True:
            for row in range(1,2):
                    letter = get_column_letter(col)
                    ws1[letter + str(row)] = result

    ws2 = wb.create_sheet(title="Ok")
    ws2["C1"] = "OK"

    nome_arquivo = '('+str(quant)+')Relatorio_Robo_'+data_inicio+'_'+hora_inicio+'_'+hora_final+'()'+result+'.xlsx'

    wb.save(nome_arquivo)

    #######################EMAIL###############################################

    # host = "smtp.gmail.com"
    # port = "587"
    # login = "andretavares16@gmail.com"
    # senha = "gmusraujapabcody"

    # server = smtplib.SMTP(host,port)

    # server.ehlo()
    # server.starttls()
    # server.login(login,senha)

    # corpo = nome_arquivo

    # email_msg = MIMEMultipart()
    # email_msg['From']=login
    # email_msg['To']=login
    # email_msg['Subject']=nome_arquivo
    # email_msg.attach(MIMEText(corpo,'plain'))

    # caminho_arquivo = "C:\\Users\\andre.tavares\\Desktop\\DEV\\ROBOS\\"+nome_arquivo
    # attachment = open(caminho_arquivo,'rb')

    # att = MIMEBase('application','octet-stream')
    # att.set_payload(attachment.read())
    # encoders.encode_base64(att)

    # att.add_header(
    #     "Content-Disposition",
    #     f"attachment; filename= {nome_arquivo}",
    # )
    # attachment.close()

    # email_msg.attach(att)

    # server.sendmail(login,login, email_msg.as_string())
    # server.quit()

    #######################TELEGRAM###################################################

    token = '5951270999:AAEPcA790iG3UJilxlVIA-WwFoQRAoSLtuc'

    chat_id = last_chat_id(token)

    print("Id do chat:",chat_id)

    send_message(token,chat_id,nome_arquivo)
    
    print('ROBO FINALIZADO !!!')
        

    





